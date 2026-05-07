from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpRequest, HttpResponse
from django.contrib import messages
from .models import EmbungModel
from .forms import EmbungForms, UploadExcelForm
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


# Create your views here.
@login_required(login_url="akun:login")
def index_embung(request):
    obj = EmbungModel.objects.all()

    context = {
        "filename": "embung",
        "data": obj,
        "form": UploadExcelForm(),
    }
    return render(request, "index_embung.html", context)


@login_required(login_url="akun:login")
def add_embung(request: HttpRequest):
    if request.method == "POST":
        form = EmbungForms(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil disimpan.")
            return redirect("embung:index")
        else:
            messages.error(request, "Form gagal disimpan.")
    else:
        form = EmbungForms()

    context = {"filename": "embung", "form": form}
    return render(request, "add_embung.html", context)


@login_required(login_url="akun:login")
def detail_embung(request: HttpRequest, id):
    object = get_object_or_404(EmbungModel, pk=id)

    context = {"filename": "embung", "object": object}

    return render(request, "detail_embung.html", context)


@login_required(login_url="akun:login")
def update_embung(request: HttpRequest, id):
    data = get_object_or_404(EmbungModel, pk=id)

    if request.method == "POST":
        form = EmbungForms(request.POST, instance=data)

        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui.")
            return redirect("embung:index")
        else:
            messages.error(request, "Gagal perbaharui data.")

    else:
        form = EmbungForms(instance=data)

    context = {
        "filename": "embung",
        "obj": data,
        "form": form,
    }

    return render(request, "edit_embung.html", context)


@login_required(login_url="akun:login")
def delete_embung(request: HttpRequest, id):
    danau = get_object_or_404(EmbungModel, pk=id)

    if request.method == "POST":
        danau.delete()
        messages.success(request, "Data berhasil dihapus.")
        return redirect("embung:index")

    return redirect("embung:index")


@login_required(login_url="akun:login")
def export_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Aset Embung"

    # Sheet Referensi Dropdown
    ref_ws = wb.create_sheet(title="Referensi")

    fields = EmbungModel._meta.get_fields()

    headers = []
    col_idx = 1

    exclude_field = [
        "id",
        "provinsi",
        "created_at",
        "updated_at",
        "created_by",
        "updated_by",
    ]

    bold_font = Font(bold=True, size=12.0)
    center_aligment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)
            ws.cell(row=1, column=col_idx, value=field.name).font = bold_font
            ws.cell(row=1, column=col_idx, value=field.name).alignment = center_aligment
            col_idx += 1

    # # NOTE: PROVINSI
    # provinsi_list = [c[0] for c in EmbungModel.PROVINSI_CHOICE]
    # for i, val in enumerate(provinsi_list, start=1):
    #     ref_ws.cell(row=i, column=1, value=val)

    # dv_provinsi = DataValidation(
    #     type="list", formula1=f"Referensi!$A$1:$A${len(provinsi_list)}"
    # )
    # ws.add_data_validation(dv_provinsi)
    # provinsi_col = headers.index("provinsi") + 1
    # dv_provinsi.add(f"{chr(64+provinsi_col)}2:{chr(64+provinsi_col)}100")

    # NOTE: KABUPATEN
    kabupaten_list = [c[0] for c in EmbungModel.KABUPATEN_CHOICE]
    for i, val in enumerate(kabupaten_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)
    dv_kabupaten = DataValidation(
        type="list", formula1=f"Referensi!$B$1:$B${len(kabupaten_list)}"
    )
    ws.add_data_validation(dv_kabupaten)

    kabupaten_col = headers.index("kabupaten") + 1
    dv_kabupaten.add(f"{chr(64+kabupaten_col)}2:{chr(64+kabupaten_col)}100")

    # NOTE: KONDISI
    kondisi_list = [c[0] for c in EmbungModel._meta.get_field("kondisi").choices]
    for i, val in enumerate(kondisi_list, start=1):
        ref_ws.cell(row=i, column=3, value=val)

    dv_kondisi = DataValidation(
        type="list", formula1=f"Referensi!$C$1:$C${len(kondisi_list)}"
    )
    ws.add_data_validation(dv_kondisi)

    kondisi_col = headers.index("kondisi") + 1
    dv_kondisi.add(f"{chr(64+kondisi_col)}2:{chr(64+kondisi_col)}100")

    repsonse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheet.sheet"
    )
    repsonse["Content-Disposition"] = 'attachment; filename="template_aset_embung.xlsx"'

    wb.save(repsonse)
    return repsonse


@login_required(login_url="akun:login")
def import_excel(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            file = request.FILES["file"]
            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                updated = 0
                skipped = 0

                for _, row in df.iterrows():

                    nama = str(row["nama"])

                    if not nama:
                        continue

                    obj = EmbungModel.objects.filter(nama=nama).first()

                    data_excel = dict(
                        nama=row["nama"],
                        kabupaten=row["kabupaten"],
                        kecamatan=row["kecamatan"],
                        desa=row["desa"],
                        latitude=row["latitude"],
                        longitude=row["longitude"],
                        tipe_konstruksi=row["tipe_konstruksi"],
                        luas_genangan=row["luas_genangan"] or None,
                        volume_tampungan=row["volume_tampungan"],
                        lebar=row["lebar"] or None,
                        panjang=row["panjang"] or None,
                        tinggi=row["tinggi"] or None,
                        irigasi=row["irigasi"] or None,
                        ternak=row["ternak"] or None,
                        air_baku=row["air_baku"] or None,
                        lainnya=row["lainnya"] or None,
                        kondisi=row["kondisi"],
                        volume_saat=row["volume_saat"] or None,
                        keterangan=row["keterangan"] or None,
                    )

                    if obj:
                        is_changed = False
                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)

                        if is_changed:
                            obj.updated_by = request.user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        EmbungModel.objects.create(
                            created_by=request.user,
                            updated_by=request.user,
                            **data_excel,
                        )
                        inserted += 1

                messages.success(
                    request, f"Insert: {inserted}, Update: {updated}, Skip: {skipped}"
                )

            except Exception as e:
                # print(str(e))
                messages.error(request, f"Error: {str(e)}")
                # raise e

            return redirect("embung:index")
    else:
        form = UploadExcelForm()

    context = {"filename": "embung", "form": form}

    return render(request, "index_danau.html", context)
