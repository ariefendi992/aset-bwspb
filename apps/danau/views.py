from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.contrib.auth.decorators import login_required
from .models import DanauModel
from .forms import DanauForms, UploadExcelForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

# from django.utils.timezone import localtime


# Create your views here.
@login_required(login_url="akun:login")
def index_danau(request):
    data = DanauModel.objects.all()
    form = UploadExcelForm()
    context = {"filename": "danau", "data": data, "form": form}

    # print(f"Local time last login ==> {localtime(request.user.last_login)}")
    # print(f"Terakhir loing ==> {request.user.last_login}")
    return render(request, "index_danau.html", context)


@login_required(login_url="akun:login")
def detail_danau(request, id):

    data = DanauModel.objects.filter(pk=id).first()
    context = {"filename": "danau", "object": data}

    return render(request, "detail_danau.html", context)


@login_required(login_url="akun:login")
def add_danau(request: HttpRequest):

    if request.method == "POST":
        form = DanauForms(request.POST)
        if form.is_valid():
            obj: DanauModel = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, "Data berhasil disimpan.")
            return redirect("danau:index")
        else:
            messages.error(request, "Form gagal disimpan.")

    else:
        form = DanauForms()

    context = {"filename": "danau", "form": form}

    return render(request, "add_danau.html", context)


@login_required(login_url="akun:login")
def update_danau(request: HttpRequest, id):
    data = get_object_or_404(DanauModel, pk=id)

    if request.method == "POST":
        form = DanauForms(request.POST, instance=data)

        if form.is_valid():
            obj: DanauModel = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            messages.success(request, "Data berhasil diperbaharui.")
            return redirect("danau:index")
        else:
            messages.error(request, "Gagal perbaharui data.")

    else:
        form = DanauForms(instance=data)

    context = {
        "filename": "danau",
        "obj": data,
        "form": form,
    }

    return render(request, "edit_danau.html", context)


@login_required(login_url="akun:login")
def delete_danau(request: HttpRequest, id):
    danau = get_object_or_404(DanauModel, pk=id)

    if request.method == "POST":
        danau.delete()
        messages.success(request, "Data berhasil dihapus.")
        return redirect("danau:index_danau")
    return redirect("danau:index")


# def peta(request: HttpRequest): ...


@login_required(login_url="akun:login")
def export_template_excel(request: HttpRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Aset Danau"

    # Sheet referensi
    ref_ws = wb.create_sheet(title="Referensi")

    fields = DanauModel._meta.get_fields()

    headers = []
    col_idx = 1

    exclude_field = ["id", "created_at", "updated_at", "created_by", "updated_by"]

    bold_font = Font(bold=True)
    center_aligment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)

            # HEADER
            cell = ws.cell(row=1, column=col_idx, value=field.name)
            cell.font = bold_font
            cell.alignment = center_aligment
            col_idx += 1

    # ---------------------------
    # PROVINSI (dropdown)
    # ---------------------------
    provinsi_list = [c[0] for c in DanauModel.PROVINSI_CHOICE]

    for i, val in enumerate(provinsi_list, start=1):
        ref_ws.cell(row=i, column=1, value=val)

    dv_prov = DataValidation(
        type="list", formula1=f"Referensi!$A$1:$A${len(provinsi_list)}"
    )
    ws.add_data_validation(dv_prov)

    prov_col = headers.index("provinsi") + 1
    dv_prov.add(f"{chr(64+prov_col)}2:{chr(64+prov_col)}100")

    # ---------------------------
    # KABUPATEN (dropdown)
    # ---------------------------
    kab_list = [c[0] for c in DanauModel.KABUPATEN_CHOICE]

    for i, val in enumerate(kab_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)

    dv_kab = DataValidation(type="list", formula1=f"Referensi!$B$1:$B${len(kab_list)}")
    ws.add_data_validation(dv_kab)

    kab_col = headers.index("kabupaten") + 1
    dv_kab.add(f"{chr(64+kab_col)}2:{chr(64+kab_col)}100")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="template_aset_danau.xlsx"'

    wb.save(response)
    return response


# def export_template_excel(request: HttpRequest):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Template Aset Danau"

#     # Ambil field dari model kecuali (id, created_by, created_at, updated_by, updated_at)
#     fields = [
#         field.name
#         for field in DanauModel._meta.fields
#         if field.name
#         not in ["id", "created_at", "updated_at", "created_by", "updated_by"]
#     ]

#     bold_font = Font(bold=True)
#     center_aligment = Alignment(horizontal="center", vertical="center")
#     # Header
#     for col_num, field in enumerate(fields, 1):
#         cell = ws.cell(row=1, column=col_num, value=field)
#         cell.font = bold_font
#         cell.alignment = center_aligment

#     # headers = [field.verbose_name for field in fields]  # jika menggunakan verbose name
#     # ws.append(fields)

#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = "attachment; filename=templates_aset_danau.xlsx"

#     wb.save(response)
#     return response


@login_required(login_url="akun:login")
def import_excel(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            file = request.FILES["file"]
            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                updated = 0
                skipped = 0

                for _, row in df.iterrows():

                    # identifikasi unique
                    nama = str(row["nama"])

                    if not nama:
                        continue
                    obj = DanauModel.objects.filter(nama=nama).first()

                    data_excel = dict(
                        nama=row["nama"],
                        provinsi=row["provinsi"],
                        kabupaten=row["kabupaten"],
                        kecamatan=row["kecamatan"],
                        desa=row["desa"] or None,
                        latitude=float(row["latitude"]),
                        longitude=float(row["longitude"]),
                        tipe_danau=row["tipe_danau"] or None,
                        luas_genangan=(
                            float(row["luas_genangan"])
                            if row["luas_genangan"] not in ["", None]
                            else None
                        ),
                        volume=(
                            float(
                                row["volume"].replace("jt", "").strip()
                                if "jt" in str(row["volume"])
                                else (row["volume"])
                            )
                            if row["volume"] not in ["", None]
                            else None
                        ),
                        irigasi=(
                            float(row["irigasi"])
                            if row["irigasi"] not in ["", None]
                            else None
                        ),
                        ternak=row["ternak"] or None,
                        air_baku=(
                            float(row["air_baku"])
                            if row["air_baku"] not in ["", None]
                            else None
                        ),
                        plta=(
                            float(row["plta"])
                            if row["plta"] not in ["", None]
                            else None
                        ),
                        volume_tampungan=(
                            float(row["volume_tampungan"].strip().split("jt")[0])
                            if row["volume_tampungan"] not in ["", None]
                            else None
                        ),
                        permasalahan=row["permasalahan"] or None,
                        keterangan=row["keterangan"] or None,
                    )

                    if obj:
                        # cek apakah ada perubahan
                        is_changed = False

                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)
                        if is_changed:
                            obj.updated_by = request.user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        DanauModel.objects.create(
                            created_by=request.user,
                            updated_by=request.user,
                            **data_excel,
                        )
                        inserted += 1

                messages.success(
                    request, f"Insert: {inserted}, Update: {updated}, Skip: {skipped}"
                )

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                # print(f"Error ==> {str(e)}")
                # print(f"❌ Error di baris {_+1}")
                # print(f"Field: latitude / longitude")
                # print(f"Value latitude: {row['latitude']}")
                # print(f"Value longitude: {row['longitude']}")
                # raise e
            return redirect("danau:index")
    else:
        form = UploadExcelForm()

    context = {"filename": "danau", "form": form}

    return render(request, "index_danau.html", context)
