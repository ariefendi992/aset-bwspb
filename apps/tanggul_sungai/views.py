from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import FileResponse, HttpRequest, Http404, HttpResponse
from .models import TanggulSungaiModel, FotoTanggulSungaiModel
from .forms import FotoTanggulSungaiForm, TanggulSungaiForm, UploadExcelForm
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


# Create your views here.
@login_required(login_url="akun:login")
def index(request):
    objects = TanggulSungaiModel.objects.order_by("-created_at").all()

    context = {"filename": "tsungai", "objects": objects, "form": UploadExcelForm()}

    return render(request, "index_tsungai.html", context)


@login_required(login_url="akun:login")
def create(request: HttpRequest):

    if request.method == "POST":
        form = TanggulSungaiForm(request.POST, request.FILES)

        if form.is_valid():
            tsungai = form.save()

            image_list = request.FILES.getlist("images")

            for image in image_list:
                FotoTanggulSungaiModel.objects.create(tanggul=tsungai, image=image)

            messages.success(request, "Data berhasil disimpan.")

            return redirect("tsungai:index")
        else:
            messages.error(request, "Form gagal disimpan.")
    else:
        form = TanggulSungaiForm()

    context = {"filename": "tsungai", "form": form}

    return render(request, "add_tsungai.html", context)


@login_required(login_url="akun:login")
def detail(request: HttpRequest, id):

    object = get_object_or_404(TanggulSungaiModel, pk=id)

    context = {"filename": "tsungai", "object": object}

    return render(request, "detail_tsungai.html", context)


@login_required(login_url="akun:login")
def update(request: HttpRequest, id):
    object = get_object_or_404(TanggulSungaiModel, pk=id)

    if request.method == "POST":
        form = TanggulSungaiForm(request.POST, request.FILES, instance=object)

        if form.is_valid():
            tsungai = form.save()
            files = request.FILES.getlist("images")

            for image in files:
                files_images = FotoTanggulSungaiModel.objects.create(
                    tanggul=tsungai, image=image
                )
                files_images.save()

            messages.success(request, "Data berhasil diperbaharui.")
            return redirect("tsungai:index")
        else:
            messages.error(request, "Gagal perbaharui data.")
    else:
        form = TanggulSungaiForm(instance=object)

    context = {"filename": "tsungai", "form": form}
    return render(request, "edit_tsungai.html", context)


@login_required(login_url="akun:login")
def delete(request: HttpRequest, id):
    object = get_object_or_404(TanggulSungaiModel, pk=id)

    if request.method == "POST":
        object.delete()

        messages.success(request, "Data berhasil dihapus.")
        return redirect("tsungai:index")

    return redirect("tsungai:index")


@login_required(login_url="akun:login")
def donwload(request, id):
    try:
        foto = FotoTanggulSungaiModel.objects.get(pk=id)
        return FileResponse(foto.image.open(), as_attachment=True)
    except:
        raise Http404("File tidak ditemukan")


@login_required(login_url="akun:login")
def export_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tanggul Sungai"

    ref_ws = wb.create_sheet(title="Referensi")

    fields = TanggulSungaiModel._meta.get_fields()
    col_idx = 1
    headers = []
    exclude_field = ["id", "created_at", "updated_at", "created_by", "updated_by"]
    bold_font = Font(bold=True, size=12.0)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)
            ws.cell(row=1, column=col_idx, value=field.name).font = bold_font
            ws.cell(row=1, column=col_idx, value=field.name).alignment = (
                center_alignment
            )

            col_idx += 1

    # NOTE: KABUPATEN
    kabupaten_list = [c[0] for c in TanggulSungaiModel.KABUPATEN_CHOICE]
    for i, val in enumerate(kabupaten_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)
    dv_kabupaten = DataValidation(
        type="list", formula1=f"Referensi!$B$1:$B${len(kabupaten_list)}"
    )
    ws.add_data_validation(dv_kabupaten)

    kabupaten_col = headers.index("kabupaten") + 1
    dv_kabupaten.add(f"{chr(64+kabupaten_col)}2:{chr(64+kabupaten_col)}100")

    # NOTE: JENIS KONSTRUKSI
    jenisKonstruksiList = [
        c[0] for c in TanggulSungaiModel._meta.get_field("jenis_konstruksi").choices
    ]
    for i, val in enumerate(jenisKonstruksiList, start=1):
        ref_ws.cell(row=i, column=3, value=val)
    dv_jenisKonstruksi = DataValidation(
        type="list", formula1=f"Referensi!$C$1:$C${len(jenisKonstruksiList)}"
    )
    ws.add_data_validation(dv_jenisKonstruksi)

    jenisKonstruksi_col = headers.index("jenis_konstruksi") + 1
    dv_jenisKonstruksi.add(
        f"{chr(64+jenisKonstruksi_col)}2:{chr(64+jenisKonstruksi_col)}100"
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="template_aset_tanggul_sungai.xlsx"'
    )

    wb.save(response)
    return response


@login_required(login_url="akun:login")
def import_template(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            file = form.files["file"]

            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                updated = 0
                skipped = 0

                for _, row in df.iterrows():
                    nama = str(row["nama"])

                    if not nama:
                        continue

                    obj = TanggulSungaiModel.objects.filter(nama=nama).first()

                    data_excel = dict(
                        nama=str(row["nama"]),
                        kabupaten=row["kabupaten"] or None,
                        latitude=row["latitude"] or None,
                        longitude=row["longitude"] or None,
                        latitude2=row["latitude2"] or None,
                        longitude2=row["longitude2"] or None,
                        lokasi_sungai=row["lokasi_sungai"] or None,
                        panjang_sungai=row["panjang_sungai"] or None,
                        nama_sungai=row["nama_sungai"] or None,
                        das=row["das"] or None,
                        jenis_konstruksi=row["jenis_konstruksi"] or None,
                        panjang_tanggul=row["panjang_tanggul"] or None,
                        tinggi_tanggul=row["tinggi_tanggul"] or None,
                        kondisi_sisi_kanan=row["kondisi_sisi_kanan"] or None,
                        kondisi_sisi_kiri=row["kondisi_sisi_kiri"] or None,
                        tahun_pembuatan=row["tahun_pembuatan"] or None,
                        status_aset=row["status_aset"] or None,
                        penamaan_bmn=row["penamaan_bmn"] or None,
                    )

                    user = request.user
                    if obj:
                        is_changed = False
                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)

                        if is_changed:
                            obj.updated_by = user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1

                    else:
                        TanggulSungaiModel.objects.create(
                            created_by=user, updated_by=user, **data_excel
                        )
                        inserted += 1
                messages.success(
                    request,
                    f"Import selesai. {inserted} data baru, {updated} data diperbaharui, {skipped} data tidak berubah.",
                )
                return redirect("tsungai:index")

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                # raise e

            return redirect("tsungai:index")
        else:
            form = UploadExcelForm()

        context = dict(filename="tsungai", form=form)

        return render(request, "index_tsungai.html", context)
