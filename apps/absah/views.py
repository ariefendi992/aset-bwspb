from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import FileResponse, Http404, HttpRequest, HttpResponse
from .models import DataAbsahModel, FotoDataAbsahModel
from .forms import AbsahForm, UploadExcelForm
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment
import pandas as pd


# Create your views here.
@login_required(login_url="akun:login")
def index(request: HttpRequest):
    objects = DataAbsahModel.objects.order_by("-created_at").all()

    context = {"filename": "absah", "objects": objects, "form": UploadExcelForm()}

    return render(request, "index_absah.html", context)


@login_required(login_url="akun:login")
def create(request: HttpRequest):

    if request.method == "POST":
        form = AbsahForm(request.POST, request.FILES)

        if form.is_valid():
            object = form.save()

            files = request.FILES.getlist("images")

            for image in files:
                foto = FotoDataAbsahModel.objects.create(absah=object, image=image)
                foto.save()
            messages.success(request, "Data berhasil disimpan.")

            return redirect("absah:index")
        else:
            messages.error(request, "Form gagal disimpan.")

    else:
        form = AbsahForm()

    context = {"filename": "absah", "form": form}

    return render(request, "add_absah.html", context)


@login_required(login_url="akun:login")
def update(request: HttpRequest, id):
    object = get_object_or_404(DataAbsahModel, pk=id)

    if request.method == "POST":
        form = AbsahForm(request.POST, request.FILES, instance=object)

        if form.is_valid():
            absah = form.save()

            files = request.FILES.getlist("images")

            for image in files:
                foto = FotoDataAbsahModel.objects.create(absah=absah, image=image)
                foto.save()

            messages.success(request, "Data berhasil disimpan.")

            return redirect("absah:index")
        else:
            messages.error(request, "Form gagal disimpan.")

    else:
        form = AbsahForm(instance=object)

    context = {"filename": "absah", "form": form, "object": object}

    return render(request, "edit_absah.html", context)


@login_required(login_url="akun:login")
def detail(request: HttpRequest, id):
    object = get_object_or_404(DataAbsahModel, pk=id)

    context = {"filename": "absah", "object": object}

    return render(request, "detail_absah.html", context)


@login_required(login_url="akun:login")
def delete(request: HttpRequest, id):
    object = get_object_or_404(DataAbsahModel, pk=id)

    if request.method == "POST":
        object.delete()
        messages.success(request, "Data berhasil dihapus.")

        return redirect("absah:index")

    return redirect("absah:index")


@login_required(login_url="akun:login")
def download_foto(request, id):
    try:
        foto = FotoDataAbsahModel.objects.get(pk=id)
        return FileResponse(foto.image.open(), as_attachment=True)
    except:
        raise Http404("File tidak ditemukan")


@login_required(login_url="akun:login")
def export_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Absah"

    ref_ws = wb.create_sheet(title="Referensi")

    fields = DataAbsahModel._meta.get_fields()
    col_idx = 1
    headers = []
    exclude_field = ["id", "created_at", "updated_at", "created_by", "updated_by"]
    bold_font = Font(bold=True, size=12.0)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)
            ws.cell(row=1, column=col_idx, value=field.name).font = bold_font
            ws.cell(row=1, column=col_idx, value=field.name).alignment = (
                center_alignment
            )

            col_idx += 1

    # NOTE: PROVINSI
    provinsi_list = [c[0] for c in DataAbsahModel.PROVINSI_CHOICE]

    for i, val in enumerate(provinsi_list, start=1):
        ref_ws.cell(row=i, column=1, value=val)

    dv_provinsi = DataValidation(
        type="list", formula1=f"Referensi!$A$1:$A${len(provinsi_list)}"
    )
    ws.add_data_validation(dv_provinsi)
    provinsi_col = headers.index("provinsi") + 1
    dv_provinsi.add(f"{chr(64+provinsi_col)}2:{chr(64+provinsi_col)}100")

    # NOTE: KABUPATEN
    kabupaten_list = [c[0] for c in DataAbsahModel.KABUPATEN_CHOICE]
    for i, val in enumerate(kabupaten_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)
    dv_kabupaten = DataValidation(
        type="list", formula1=f"Referensi!$B$1:$B${len(kabupaten_list)}"
    )
    ws.add_data_validation(dv_kabupaten)

    kabupaten_col = headers.index("kabupaten") + 1
    dv_kabupaten.add(f"{chr(64+kabupaten_col)}2:{chr(64+kabupaten_col)}100")

    repsonse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheet.sheet"
    )
    repsonse["Content-Disposition"] = 'attachment; filename="template_aset_absah.xlsx"'

    wb.save(repsonse)
    return repsonse


@login_required(login_url="akun:login")
def import_template(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            file = form.files["file"]

            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                updated = 0
                skipped = 0

                for _, row in df.iterrows():
                    nama = str(row["nama"])

                    if not nama:
                        continue
                    obj = DataAbsahModel.objects.filter(nama=nama).first()

                    data_excel = dict(
                        nama=nama,
                        provinsi=str(row["provinsi"]) or None,
                        kabupaten=str(row["kabupaten"]) or None,
                        kecamatan=str(row["kecamatan"]) or None,
                        desa=str(row["desa"]) or None,
                        latitude=row["latitude"] or None,
                        longitude=row["longitude"] or None,
                        volume=row["volume"] or None,
                        panjang=row["panjang"] or None,
                        tinggi=row["tinggi"] or None,
                        lebar=row["lebar"] or None,
                        tahun_mulai=row["tahun_mulai"] or None,
                        tahun_selesai=row["tahun_selesai"] or None,
                        pengelola=row["pengelola"] or None,
                        keterangan=row["keterangan"] or None,
                        manfaat=row["manfaat"] or None,
                    )

                    if obj:
                        is_changed = False
                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)

                        if is_changed:
                            obj.updated_by = request.user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        DataAbsahModel.objects.create(
                            created_by=request.user,
                            updated_by=request.user,
                            **data_excel,
                        )
                        inserted += 1

                messages.success(
                    request,
                    f"Import selesai. {inserted} data baru, {updated} data diperbaharui, {skipped} data tidak berubah.",
                )
                return redirect("absah:index")

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

                # raise e

            return redirect("absah:index")

        else:
            form = UploadExcelForm()

        context = {"filename": "absah", "form": form}

        return render(request, "index_absah.html", context)
