from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, Http404, HttpRequest, HttpResponse
from django.contrib import messages
from apps.asets.models import CheckDamModel, FotoCheckDamModel
from apps.asets.forms import CheckDamForm, UploadExcelFormCheckDam
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


# Create your views here.
@login_required(login_url="akun:login")
def index(request):

    obj = CheckDamModel.objects.all().prefetch_related("dokumentasi")

    context = {
        "filename": "checkdam",
        "objects": obj,
        "form": UploadExcelFormCheckDam(),
    }

    return render(request, "checkdam/index_checkdam.html", context)


@login_required(login_url="akun:login")
def create(request: HttpRequest):

    if request.method == "POST":
        form = CheckDamForm(request.POST, request.FILES)

        if form.is_valid():
            checkdam = form.save()
            image_list = request.FILES.getlist("images")

            for image in image_list:
                FotoCheckDamModel.objects.create(checkdam=checkdam, image=image)

            messages.success(request, "Data berhasil disimpan.")

            return redirect("checkdam:index")
        else:
            messages.error(request, "Form gagal disimpan.")
    else:
        form = CheckDamForm()

    context = {"filename": "checkdam", "form": form}

    return render(request, "checkdam/add_checkdam.html", context)


@login_required(login_url="akun:login")
def detail(request: HttpRequest, obj_id):
    object = get_object_or_404(CheckDamModel, pk=obj_id)

    context = {"filename": "checkdam", "object": object}

    return render(request, "checkdam/detail_checkdam.html", context)


@login_required(login_url="akun:login")
def update(request: HttpRequest, obj_id):

    checkdam = get_object_or_404(CheckDamModel, pk=obj_id)

    if request.method == "POST":
        form = CheckDamForm(request.POST, request.FILES, instance=checkdam)
        if form.is_valid():
            checkdam_ = form.save()

            # ambil multiple file
            files = request.FILES.getlist("images")

            for file in files:
                FotoCheckDamModel.objects.create(checkdam=checkdam_, image=file)

            messages.success(request, "Data berhasil diperbaharui.")
            return redirect("checkdam:index")
        else:
            messages.error(request, "Gagal perbaharui data.")
    else:
        form = CheckDamForm(instance=checkdam)

    fotos = checkdam.dokumentasi.all()

    context = {"filename": "checkdam", "form": form, "fotos": fotos}
    return render(request, "checkdam/edit_checkdam.html", context)


@login_required(login_url="akun:login")
def delete(request: HttpRequest, obj_id):
    checkdam = get_object_or_404(CheckDamModel, pk=obj_id)

    if request.method == "POST":
        checkdam.delete()
        messages.success(request, "Data berhasil dihapus.")
        return redirect("checkdam:index")

    return redirect("checkdam:index")


@login_required(login_url="akun:login")
def download_foto(request, foto_id):
    try:
        foto = FotoCheckDamModel.objects.get(id=foto_id)
        return FileResponse(foto.image.open(), as_attachment=True)
    except:
        raise Http404("File tidak ditemukan")


@login_required(login_url="akun:login")
def export_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Check Dam"

    ref_ws = wb.create_sheet(title="Referensi")

    fields = CheckDamModel._meta.get_fields()
    col_idx = 1
    headers = []
    exclude_field = ["id", "created_at", "updated_at", "created_by", "updated_by"]
    bold_font = Font(bold=True, size=12.0)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)
            ws.cell(row=1, column=col_idx, value=field.name).font = bold_font
            ws.cell(row=1, column=col_idx, value=field.name).alignment = (
                center_alignment
            )

            col_idx += 1

    provinsi_list = [c[0] for c in CheckDamModel.PROVINSI_CHOICE]

    for i, val in enumerate(provinsi_list, start=1):
        ref_ws.cell(row=i, column=1, value=val)

    dv_provinsi = DataValidation(
        type="list", formula1=f"Referensi!$A$1:$A${len(provinsi_list)}"
    )
    ws.add_data_validation(dv_provinsi)
    provinsi_col = headers.index("provinsi") + 1
    dv_provinsi.add(f"{chr(64+provinsi_col)}2:{chr(64+provinsi_col)}100")

    # NOTE: KABUPATEN
    kabupaten_list = [c[0] for c in CheckDamModel.KABUPATEN_CHOICE]
    for i, val in enumerate(kabupaten_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)
    dv_kabupaten = DataValidation(
        type="list", formula1=f"Referensi!$B$1:$B${len(kabupaten_list)}"
    )
    ws.add_data_validation(dv_kabupaten)

    kabupaten_col = headers.index("kabupaten") + 1
    dv_kabupaten.add(f"{chr(64+kabupaten_col)}2:{chr(64+kabupaten_col)}100")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="template_aset_checkdam.xlsx"'
    )

    wb.save(response)
    return response


@login_required(login_url="akun:login")
def import_template(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelFormCheckDam(request.POST, request.FILES)

        if form.is_valid():
            file = request.FILES["file"]

            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                updated = 0
                skipped = 0

                for _, row in df.iterrows():
                    nama = str(row["nama"])

                    if not nama:
                        continue
                    obj = CheckDamModel.objects.filter(nama=nama).first()

                    data_excel = dict(
                        nama=nama,
                        provinsi=str(row["provinsi"]) or None,
                        kabupaten=str(row["kabupaten"]) or None,
                        kecamatan=str(row["kecamatan"]) or None,
                        desa=str(row["desa"]) or None,
                        volume_sabo=row["volume_sabo"] or None,
                        panjang=row["panjang"] or None,
                        tinggi=row["tinggi"] or None,
                        lebar=row["lebar"] or None,
                        tahun_mulai=row["tahun_mulai"] or None,
                        tahun_selesai=row["tahun_selesai"] or None,
                        pengelola=row["pengelola"] or None,
                        keterangan=row["keterangan"] or None,
                        irigasi=row["irigasi"] or None,
                        lain=row["lain"] or None,
                        latitude=row["latitude"] or None,
                        longitude=row["longitude"] or None,
                        latitude2=row["latitude2"] or None,
                        longitude2=row["longitude2"] or None,
                        status_aset=row["status_aset"] or None,
                        penamaan_bmn=row["penamaan_bmn"] or None,
                    )

                    if obj:
                        is_changed = False
                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)

                        if is_changed:
                            obj.updated_by = request.user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        CheckDamModel.objects.create(
                            created_by=request.user,
                            updated_by=request.user,
                            **data_excel,
                        )
                        inserted += 1

                messages.success(
                    request,
                    f"Import selesai. {inserted} data baru, {updated} data diperbaharui, {skipped} data tidak berubah.",
                )
                return redirect("checkdam:index")

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

                # raise e

            return redirect("checkdam:index")

        else:
            form = UploadExcelFormCheckDam()

        context = {"filename": "checkdam", "form": form}

        return render(request, "checkdam/index_checkdam.html", context)
