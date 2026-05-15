from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpRequest, Http404, HttpResponse
from django.contrib import messages
from apps.asets.models import PengamanPantaiModel
from apps.asets.forms import PengamanPantaiForm, UploadExcelFormPantai
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


# Create your views here.
@login_required(login_url="akun:login")
def index(request: HttpRequest):

    objects = PengamanPantaiModel.objects.order_by("-created_at").all()

    context = {
        "filename": "ppantai",
        "objects": objects,
        "form": UploadExcelFormPantai(),
    }

    return render(request, "pantai/index_ppantai.html", context)


@login_required(login_url="akun:login")
def create(request: HttpRequest):

    if request.method == "POST":
        form = PengamanPantaiForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil disimpan.")

            return redirect("ppantai:index")
        else:
            messages.error(request, "Form gagal disimpan.")
    else:
        form = PengamanPantaiForm()

    context = {"filename": "ppantai", "form": form}
    return render(request, "pantai/add_ppantai.html", context)


@login_required(login_url="akun:login")
def detail(request: HttpRequest, obj_id):

    object = get_object_or_404(PengamanPantaiModel, pk=obj_id)

    print(f"Latitude Pantai ==> {object.latitude}")

    context = {"filename": "ppantai", "object": object}
    return render(request, "pantai/detail_ppantai.html", context)


@login_required(login_url="akun:login")
def update(request: HttpRequest, obj_id):
    object = get_object_or_404(PengamanPantaiModel, pk=obj_id)

    if request.method == "POST":
        form = PengamanPantaiForm(request.POST, instance=object)
        if form.is_valid():
            form.save()

            messages.success(request, "Data berhasil diperbaharui.")

            return redirect("ppantai:index")

        else:
            messages.error(request, "Gagal perbaharui data.")

    else:
        form = PengamanPantaiForm(instance=object)

    context = {"filename": "ppantai", "form": form, "obj": object}

    return render(request, "pantai/edit_ppantai.html", context)


@login_required(login_url="akun:login")
def delete(request: HttpRequest, obj_id):
    object = get_object_or_404(PengamanPantaiModel, pk=obj_id)

    if request.method == "POST":
        object.delete()
        messages.success(request, "Data berhasil dihapus.")

        return redirect("ppantai:index")
    return redirect("ppantai:index")


@login_required(login_url="akun:login")
def export_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pengaman Pantai"

    ref_ws = wb.create_sheet(title="Referensi")

    fields = PengamanPantaiModel._meta.get_fields()
    col_idx = 1
    headers = []
    exclude_field = ["id", "created_at", "updated_at", "created_by", "updated_by"]
    bold_font = Font(bold=True, size=12.0)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)
            ws.cell(row=1, column=col_idx, value=field.name).font = bold_font
            ws.cell(row=1, column=col_idx, value=field.name).alignment = (
                center_alignment
            )

            col_idx += 1

    provinsi_list = [c[0] for c in PengamanPantaiModel.PROVINSI_CHOICE]

    for i, val in enumerate(provinsi_list, start=1):
        ref_ws.cell(row=i, column=1, value=val)

    dv_provinsi = DataValidation(
        type="list", formula1=f"Referensi!$A$1:$A${len(provinsi_list)}"
    )
    ws.add_data_validation(dv_provinsi)
    provinsi_col = headers.index("provinsi") + 1
    dv_provinsi.add(f"{chr(64+provinsi_col)}2:{chr(64+provinsi_col)}100")

    # NOTE: KABUPATEN
    kabupaten_list = [c[0] for c in PengamanPantaiModel.KABUPATEN_CHOICE]
    for i, val in enumerate(kabupaten_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)
    dv_kabupaten = DataValidation(
        type="list", formula1=f"Referensi!$B$1:$B${len(kabupaten_list)}"
    )
    ws.add_data_validation(dv_kabupaten)

    kabupaten_col = headers.index("kabupaten") + 1
    dv_kabupaten.add(f"{chr(64+kabupaten_col)}2:{chr(64+kabupaten_col)}100")

    # NOTE: JENIS BANGUNAN
    jenisBangunan_list = [c[0] for c in PengamanPantaiModel.JENIS_BANGUNAN_CHOICE]
    for i, val in enumerate(jenisBangunan_list, start=1):
        ref_ws.cell(row=i, column=3, value=val)
    dv_jenisBangunan = DataValidation(
        type="list", formula1=f"Referensi!$C$1:$C${len(jenisBangunan_list)}"
    )
    ws.add_data_validation(dv_jenisBangunan)

    jenisBangunan_col = headers.index("jenis_bangunan") + 1
    dv_jenisBangunan.add(f"{chr(64+jenisBangunan_col)}2:{chr(64+jenisBangunan_col)}100")

    # NOTE: MATERIAL
    material_list = [
        c[0] for c in PengamanPantaiModel._meta.get_field("material").choices
    ]
    for i, val in enumerate(material_list, start=1):
        ref_ws.cell(row=i, column=4, value=val)
    dv_material = DataValidation(
        type="list", formula1=f"Referensi!$D$1:$D${len(material_list)}"
    )
    ws.add_data_validation(dv_material)

    material_col = headers.index("material") + 1
    dv_material.add(f"{chr(64+material_col)}2:{chr(64+material_col)}100")

    # NOTE: STRUKTUR
    struktur_list = [
        c[0] for c in PengamanPantaiModel._meta.get_field("struktur").choices
    ]
    for i, val in enumerate(struktur_list, start=1):
        ref_ws.cell(row=i, column=5, value=val)
    dv_struktur = DataValidation(
        type="list", formula1=f"Referensi!$E$1:$E${len(struktur_list)}"
    )
    ws.add_data_validation(dv_struktur)

    struktur_col = headers.index("material") + 1
    dv_struktur.add(f"{chr(64+struktur_col)}2:{chr(64+struktur_col)}100")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="template_aset_pengaman_pantai.xlsx"'
    )

    wb.save(response)
    return response


@login_required(login_url="akun:login")
def import_template(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelFormPantai(request.POST, request.FILES)

        if form.is_valid():
            file = request.FILES["file"]
            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                updated = 0
                skipped = 0

                for _, row in df.iterrows():
                    nama = str(row["nama"])
                    panjang = row["panjang"] or None
                    lebar_puncak = row["lebar_puncak"] or None

                    # print(f"{_} => {nama}")

                    if not nama and not panjang and not lebar_puncak:
                        continue

                    obj = PengamanPantaiModel.objects.filter(
                        nama=nama, panjang=panjang, lebar_puncak=lebar_puncak
                    ).first()

                    data_excel = dict(
                        nama=nama,
                        provinsi=str(row["provinsi"]) or None,
                        kabupaten=str(row["kabupaten"]) or None,
                        kecamatan=str(row["kecamatan"]) or None,
                        desa=str(row["desa"]) or None,
                        latitude=row["latitude"] or None,
                        longitude=row["longitude"] or None,
                        kewenangan=row["kewenangan"] or None,
                        jenis_bangunan=row["jenis_bangunan"] or None,
                        panjang=row["panjang"] or None,
                        lebar_puncak=row["lebar_puncak"] or None,
                        kemiringan=row["kemiringan"] or None,
                        material=row["material"] or None,
                        struktur=row["struktur"] or None,
                        tahun_mulai=row["tahun_mulai"] or None,
                        tahun_selesai=row["tahun_selesai"] or None,
                        dibangun_oleh=row["dibangun_oleh"] or None,
                        pelindung=row["pelindung"] or None,
                        lain_lain=row["lain_lain"] or None,
                        latitude2=row["latitude2"] or None,
                        longitude2=row["longitude2"] or None,
                    )

                    # print(f"Nomor: {_}, data excel ==> {data_excel['nama']}")

                    user = request.user

                    if obj:
                        is_changed = False
                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)

                        if is_changed:
                            obj.updated_by = user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1

                    else:
                        PengamanPantaiModel.objects.create(
                            created_by=user,
                            updated_by=user,
                            **data_excel,
                        )

                        inserted += 1

                messages.success(
                    request,
                    f"Import selesai. {inserted} data baru, {updated} data diperbaharui, {skipped} data tidak berubah.",
                )
                return redirect("ppantai:index")

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                # raise e

            return redirect("ppantai:index")

        else:
            form = UploadExcelFormPantai()
            messages.error(request, message=f"terjadi kesalahan upload file")
            # raise

        context = dict(filename="ppantai", form=form)

        return render(request, "pantai/index_ppantai.html", context)
