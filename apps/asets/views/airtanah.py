from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpRequest, HttpResponse
from django.contrib import messages
from apps.asets.models import SumurAirTanahModel
from apps.asets.forms import SumurAirTanahForm, UploadExcelFormAirTanah
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


# Create your views here.
@login_required(login_url="akun:login")
def index(request):
    objects = SumurAirTanahModel.objects.order_by("-created_at").all()

    context = {
        "filename": "sumur",
        "objects": objects,
        "form": UploadExcelFormAirTanah(),
    }
    return render(request, "airtanah/index_sumur.html", context)


def create(request: HttpRequest):

    if request.method == "POST":
        form = SumurAirTanahForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil disimpan.")
            return redirect("sumur:index")
        else:
            messages.error(request, "Form gagal disimpan.")
    else:
        form = SumurAirTanahForm()

    context = {
        "filename": "sumur",
        "form": form,
    }

    return render(request, "airtanah/add_sumur.html", context)


def detail(request: HttpRequest, id):
    object = get_object_or_404(SumurAirTanahModel, pk=id)

    context = {"filename": "sumur", "object": object}

    return render(request, "airtanah/detail_sumur.html", context)


def update(request: HttpRequest, id):
    object = get_object_or_404(SumurAirTanahModel, pk=id)
    if request.method == "POST":
        form = SumurAirTanahForm(request.POST, instance=object)

        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui.")

            return redirect("sumur:index")
        else:
            messages.error(request, "Gagal perbaharui data.")
    else:
        form = SumurAirTanahForm(instance=object)

    context = {"filename": "sumur", "form": form, "object": object}

    return render(request, "airtanah/edit_sumur.html", context)


def delete(request: HttpRequest, id):
    object = get_object_or_404(SumurAirTanahModel, pk=id)
    if request.method == "POST":

        object.delete()

        messages.success(request, "Data berhasil dihapus.")

        return redirect("sumur:index")

    return redirect("sumur:index")


@login_required(login_url="akun:login")
def export_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sumur Air Tanah"

    # Sheet Referensi Dropdown
    ref_ws = wb.create_sheet(title="Referensi")

    fields = SumurAirTanahModel._meta.get_fields()

    headers = []
    col_idx = 1

    exclude_field = [
        "id",
        "provinsi",
        "created_at",
        "updated_at",
        "created_by",
        "updated_by",
    ]

    bold_font = Font(bold=True, size=12.0)
    center_aligment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)
            ws.cell(row=1, column=col_idx, value=field.name).font = bold_font
            ws.cell(row=1, column=col_idx, value=field.name).alignment = center_aligment
            col_idx += 1

    # # NOTE: PROVINSI
    # provinsi_list = [c[0] for c in EmbungModel.PROVINSI_CHOICE]
    # for i, val in enumerate(provinsi_list, start=1):
    #     ref_ws.cell(row=i, column=1, value=val)

    # dv_provinsi = DataValidation(
    #     type="list", formula1=f"Referensi!$A$1:$A${len(provinsi_list)}"
    # )
    # ws.add_data_validation(dv_provinsi)
    # provinsi_col = headers.index("provinsi") + 1
    # dv_provinsi.add(f"{chr(64+provinsi_col)}2:{chr(64+provinsi_col)}100")

    # NOTE: KABUPATEN
    kabupaten_list = [c[0] for c in SumurAirTanahModel.KABUPATEN_CHOICE]
    for i, val in enumerate(kabupaten_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)
    dv_kabupaten = DataValidation(
        type="list", formula1=f"Referensi!$B$1:$B${len(kabupaten_list)}"
    )
    ws.add_data_validation(dv_kabupaten)

    kabupaten_col = headers.index("kabupaten") + 1
    dv_kabupaten.add(f"{chr(64+kabupaten_col)}2:{chr(64+kabupaten_col)}100")

    repsonse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheet.sheet"
    )
    repsonse["Content-Disposition"] = (
        'attachment; filename="template_aset_air_tanah.xlsx"'
    )

    wb.save(repsonse)
    return repsonse


def import_excel(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelFormAirTanah(request.POST, request.FILES)

        if form.is_valid():
            file = form.files["file"]
            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                updated = 0
                skipped = 0

                for _, row in df.iterrows():

                    nama = str(row["nama"])

                    if not nama:
                        continue

                    obj = SumurAirTanahModel.objects.filter(nama=nama).first()

                    data_excel = dict(
                        nama=row["nama"],
                        kabupaten=row["kabupaten"] or None,
                        kecamatan=row["kecamatan"] or None,
                        desa=row["desa"] or None,
                        latitude=row["latitude"] or None,
                        longitude=row["longitude"] or None,
                        jenis_sumur=row["jenis_sumur"] or None,
                        kedalaman_sumur=row["kedalaman_sumur"] or None,
                        jenis_pompa=row["jenis_pompa"] or None,
                        debit_pompa=row["debit_pompa"] or None,
                        tahun_pembangunan=row["tahun_pembangunan"] or None,
                        das=row["das"] or None,
                        ws=row["ws"] or None,
                        keterangan=row["keterangan"] or None,
                    )

                    if obj:
                        is_changed = False
                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)

                        if is_changed:
                            obj.updated_by = request.user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        SumurAirTanahModel.objects.create(
                            created_by=request.user,
                            updated_by=request.user,
                            **data_excel,
                        )
                        inserted += 1

                messages.success(
                    request, f"Insert: {inserted}, Update: {updated}, Skip: {skipped}"
                )

            except Exception as e:
                # print(str(e))
                messages.error(request, f"Error: {str(e)}")
                # raise e

            return redirect("sumur:index")
    else:
        form = UploadExcelFormAirTanah()

    context = {"filename": "sumur", "form": form}

    return render(request, "airtanah/index_sumur.html", context)
