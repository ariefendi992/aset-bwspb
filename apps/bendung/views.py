from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import FileResponse, Http404, HttpRequest, HttpResponse
from apps.bendung.models import FotoBendungModel, BendungModel
from .forms import BendungForms, FotoBendungFormSet, UploadExcelForm
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


# Create your views here.
@login_required(login_url="akun:login")
def index_bendung(request):

    obj = BendungModel.objects.all().prefetch_related("dokumentasi")
    context = {
        "filename": "bendung",
        "objects": obj,
        "form": UploadExcelForm,
    }
    return render(request, "index_bendung.html", context)


@login_required(login_url="akun:login")
def add_bendung(request: HttpRequest):
    if request.method == "POST":
        # formset = FotoBendungFormSet(request.POST, request.FILES)
        form = BendungForms(request.POST, request.FILES)

        if form.is_valid():
            bendung = form.save()
            image_list = request.FILES.getlist("images")  # ✅ FIX DI SINI

            for image in image_list:
                FotoBendungModel.objects.create(bendung=bendung, image=image)
            # formset.instance = bendung
            # formset.save()
            messages.success(request, "Data berhasil disimpan.")
            return redirect("bendung:index_bendung")
        else:
            messages.error(request, "Form gagal disimpan.")

    else:
        form = BendungForms()
        # formset = FotoBendungFormSet()
    context = {
        "filename": "bendung",
        "form": form,
        # "formset": formset,
    }

    return render(request, "add_bendung.html", context)


@login_required(login_url="akun:login")
def update_bendung(request: HttpRequest, id):
    object = get_object_or_404(BendungModel, pk=id)
    if request.method == "POST":
        # formset = FotoBendungFormSet(request.POST, request.FILES)
        form = BendungForms(
            request.POST, request.FILES, instance=object
        )  # ✅ FIX DI SINI

        if form.is_valid():
            bendung = form.save()
            image_list = request.FILES.getlist("images")  # ✅ FIX DI SINI

            for image in image_list:
                FotoBendungModel.objects.create(bendung=bendung, image=image)
            # formset.instance = bendung
            # formset.save()
            messages.success(request, "Data berhasil diperbaharui.")
            return redirect("bendung:index_bendung")
        else:
            messages.error(request, "Gagal perbaharui data.")

    else:
        form = BendungForms(instance=object)
        # formset = FotoBendungFormSet()
    context = {
        "filename": "bendung",
        "form": form,
        # "formset": formset,
    }

    return render(request, "edit_bendung.html", context)


@login_required(login_url="akun:login")
def download_foto(request, foto_id):
    try:
        foto = FotoBendungModel.objects.get(id=foto_id)
        return FileResponse(foto.image.open(), as_attachment=True)
    except:
        raise Http404("File tidak ditemukan")


@login_required(login_url="akun:login")
def delete_data(request: HttpRequest, id):
    danau = get_object_or_404(BendungModel, pk=id)

    if request.method == "POST":
        danau.delete()
        messages.success(request, "Data berhasil dihapus.")
        return redirect("bendung:index_bendung")
    return redirect("bendung:index_bendung")


@login_required(login_url="akun:login")
def detail_bendung(request: HttpRequest, id):

    object = get_object_or_404(BendungModel, pk=id)

    context = {
        "filename": "bendung",
        "object": object,
    }

    return render(request, "detail_bendung.html", context)


@login_required(login_url="akun:login")
def export_template(request: HttpRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Aset Bendung"

    ref_ws = wb.create_sheet(title="Referensi")

    fields = BendungModel._meta.get_fields()
    col_idx = 1
    headers = []
    exclude_field = ["id", "created_at", "updated_at", "created_by", "updated_by"]
    bold_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for field in fields:
        if (
            field.concrete
            and not field.many_to_many
            and field.name not in exclude_field
        ):
            headers.append(field.name)

            ws.cell(row=1, column=col_idx, value=field.name).font = bold_font
            ws.cell(row=1, column=col_idx, value=field.name).alignment = (
                center_alignment
            )
            col_idx += 1

    provinsi_list = [c[0] for c in BendungModel.PROVINSI_CHOICE]

    for i, val in enumerate(provinsi_list, start=1):
        ref_ws.cell(row=i, column=1, value=val)

    dv_provinsi = DataValidation(
        type="list", formula1=f"Referensi!$A$1:$A${len(provinsi_list)}"
    )
    ws.add_data_validation(dv_provinsi)
    provinsi_col = headers.index("provinsi") + 1
    dv_provinsi.add(f"{chr(64+provinsi_col)}2:{chr(64+provinsi_col)}100")

    # NOTE: KABUPATEN
    kabupaten_list = [c[0] for c in BendungModel.KABUPATEN_CHOICE]
    for i, val in enumerate(kabupaten_list, start=1):
        ref_ws.cell(row=i, column=2, value=val)
    dv_kabupaten = DataValidation(
        type="list", formula1=f"Referensi!$B$1:$B${len(kabupaten_list)}"
    )
    ws.add_data_validation(dv_kabupaten)

    kabupaten_col = headers.index("kabupaten") + 1
    dv_kabupaten.add(f"{chr(64+kabupaten_col)}2:{chr(64+kabupaten_col)}100")

    # NOTE: KONDISI
    kondisi_list = [c[0] for c in BendungModel.KONDISI_CHOICE]
    for i, val in enumerate(kondisi_list, start=1):
        ref_ws.cell(row=i, column=3, value=val)

    dv_kondisi = DataValidation(
        type="list", formula1=f"Referensi!$C$1:$C${len(kondisi_list)}"
    )
    ws.add_data_validation(dv_kondisi)

    kondisi_col = headers.index("kondisi") + 1
    dv_kondisi.add(f"{chr(64+kondisi_col)}2:{chr(64+kondisi_col)}100")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="template_aset_bendung.xlsx"'
    )

    wb.save(response)
    return response


@login_required(login_url="akun:login")
def import_excel(request: HttpRequest):
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            file = request.FILES["file"]
            try:
                df = pd.read_excel(file)
                df.columns = df.columns.str.strip().str.lower()
                df = df.fillna("")

                inserted = 0
                skipped = 0
                updated = 0

                for _, row in df.iterrows():

                    nama = str(row["nama"])

                    if not nama:
                        continue

                    obj = BendungModel.objects.filter(nama=nama).first()

                    data_excel = dict(
                        nama=nama,
                        provinsi=str(row["provinsi"]) or None,
                        kabupaten=str(row["kabupaten"]) or None,
                        kecamatan=str(row["kecamatan"]) or None,
                        desa=str(row["desa"]) or None,
                        sungai=str(row["sungai"]) or None,
                        jenis_bendung=row["jenis_bendung"] or None,
                        tinggi=row["tinggi"] or None,
                        lebar=row["lebar"] or None,
                        debit_intake_musim_hujan=row["debit_intake_musim_hujan"]
                        or None,
                        debit_intake_musim_kemarau=row["debit_intake_musim_kemarau"]
                        or None,
                        tahun_mulai=row["tahun_mulai"] or None,
                        tahun_rehab=row["tahun_rehab"] or None,
                        kondisi=str(row["kondisi"]) or None,
                        irigasi=row["irigasi"] or None,
                        lain_lain=str(row["lain_lain"]) or None,
                        latitude1=row["latitude1"] or None,
                        longitude1=row["longitude1"] or None,
                        latitude2=row["latitude2"] or None,
                        longitude2=row["longitude2"] or None,
                        status_aset=str(row["status_aset"]) or None,
                        penamaan_bmn=str(row["penamaan_bmn"]) or None,
                    )

                    if obj:
                        is_changed = False
                        for field, value in data_excel.items():
                            if getattr(obj, field) != value:
                                is_changed = True
                                setattr(obj, field, value)
                        if is_changed:
                            obj.updated_by = request.user
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        BendungModel.objects.create(
                            created_by=request.user,
                            updated_by=request.user,
                            **data_excel,
                        )
                        inserted += 1
                messages.success(
                    request,
                    f"Import selesai. {inserted} data baru, {updated} data diperbaharui, {skipped} data tidak berubah.",
                )
                return redirect("bendung:index")

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                # raise e
            return redirect("bendung:index")
    else:
        form = UploadExcelForm()

    context = {"filename": "bendung", "form": form}
    return render(request, "index_bendung.html", context)
